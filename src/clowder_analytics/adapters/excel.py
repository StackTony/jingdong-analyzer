"""F002 P1: Excel Adapter（spec §3.3）

读 .xlsx / .xls 文件，支持多 sheet（默认第一个）。
输出 Dataset 对象含 schema_fingerprint。

G1 大数据量优化：支持 max_rows 采样加载（openpyxl read_only 流式 + 截断），
避免大 xlsx 全量加载 OOM。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

from clowder_analytics.adapters.base import (
    ColumnSpec,
    Dataset,
    DataSourceAdapter,
    compute_fingerprint,
)


class ExcelAdapter(DataSourceAdapter):
    """读 Excel 文件为 Dataset"""

    def load(self, source_desc: dict[str, Any]) -> Dataset:
        path = Path(source_desc["path"])
        sheet = source_desc.get("sheet", 0)  # 0 = 第一个 sheet
        max_rows = source_desc.get("max_rows")

        # 先拿全量行数（openpyxl read_only 流式，不全部加载到 DataFrame）
        full_row_count = _count_excel_rows(path, sheet)
        sampled = False

        if max_rows is not None:
            # openpyxl read_only 模式只读前 max_rows+1 行（+1 header）
            df = pd.read_excel(path, sheet_name=sheet, nrows=max_rows)
            sampled = True
        else:
            df = pd.read_excel(path, sheet_name=sheet)

        column_hints = source_desc.get("column_hints")
        fp = compute_fingerprint(df, column_hints)
        columns = [
            ColumnSpec(name=c, dtype=str(df[c].dtype))
            for c in df.columns
        ]
        return Dataset(
            df=df,
            schema_fingerprint=fp,
            metadata={
                "path": str(path),
                "sheet": str(sheet),
                "row_count": len(df),
                "sampled": sampled,
                "full_row_count": full_row_count,
            },
            source_type="excel",
            columns=columns,
        )

    def supported_types(self) -> list[str]:
        return ["excel"]


def _count_excel_rows(path: Path, sheet) -> int:
    """数 Excel sheet 总行数（不含 header），openpyxl read_only 流式"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet if isinstance(sheet, int) else 0]
        count = ws.max_row - 1  # 减 header
        wb.close()
        return max(count, 0)
    except Exception:
        return 0
